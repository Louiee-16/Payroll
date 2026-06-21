import calendar as cal
import datetime
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from employees.models import Employees
from LEAVES.models import LeaveRequest
from LOANS.models import loans
from LOGS.models import Attendance
from user.decorators import staff_required

from user.models import log_action
from .models import PayrollRun, PayrollPeriod, Contribution


@staff_required
def PAYROLL(request):
    frequency = request.GET.get('frequency', 'monthly')
    today     = datetime.date.today()

    # ── Period boundaries ────────────────────────────────────────────────
    if frequency == 'weekly':
        period_start = today - datetime.timedelta(days=today.weekday())  # Monday
        period_end   = period_start + datetime.timedelta(days=6)         # Sunday
        divisor      = Decimal('4')
        freq_label   = 'Weekly'
        period_label = f"{period_start.strftime('%b %d')} - {period_end.strftime('%b %d, %Y')}"

    elif frequency == 'biweekly':
        if today.day <= 15:
            period_start = today.replace(day=1)
            period_end   = today.replace(day=15)
        else:
            last_day     = cal.monthrange(today.year, today.month)[1]
            period_start = today.replace(day=16)
            period_end   = today.replace(day=last_day)
        divisor      = Decimal('2')
        freq_label   = 'Every 2 Weeks'
        period_label = f"{period_start.strftime('%b %d')} - {period_end.strftime('%b %d, %Y')}"

    else:  # monthly
        last_day     = cal.monthrange(today.year, today.month)[1]
        period_start = today.replace(day=1)
        period_end   = today.replace(day=last_day)
        divisor      = Decimal('1')
        freq_label   = 'Monthly'
        period_label = today.strftime('%B %Y')

    # Only count up to today so absent/days-worked figures are meaningful
    effective_end = min(period_end, today)

    total_working_days = sum(
        1 for i in range((effective_end - period_start).days + 1)
        if (period_start + datetime.timedelta(days=i)).weekday() < 5
    )

    # ── Bulk-fetch attendance (avoids N+1) ──────────────────────────────
    attendance_by_emp = {}
    for att in Attendance.objects.filter(
        date__range=(period_start, effective_end)
    ).select_related('employee'):
        attendance_by_emp.setdefault(att.employee_id, []).append(att)

    employees = list(Employees.objects.prefetch_related('employee_loans'))

    total_gross          = Decimal('0')
    total_deductions_all = Decimal('0')
    total_net_all        = Decimal('0')

    for emp in employees:
        monthly      = Decimal(str(emp.basic_pay))
        period_gross = (monthly / divisor).quantize(Decimal('0.01'))

        emp_logs    = attendance_by_emp.get(emp.id, [])
        days_worked = len(emp_logs)
        absent_days = max(0, total_working_days - days_worked)

        late_mins = sum(
            late_tracker(att.time_in, emp.duty_in)
            for att in emp_logs if att.time_in
        )

        leave_count = LeaveRequest.objects.filter(
            employee=emp,
            status='APPROVED',
            start_date__lte=effective_end,
            end_date__gte=period_start,
        ).count()

        # Overtime
        ot_hours, ot_pay = calculate_overtime(emp_logs, emp.duty_out, monthly)

        # Adjustments for this period
        adj_earnings, adj_deductions = get_adjustment_totals(emp, period_start)

        # 13th month accrual
        thirteenth = calculate_13th_month(monthly, emp.date_hired)

        # Gross = base period pay + OT + adjustment earnings
        period_gross = (monthly / divisor + ot_pay + adj_earnings).quantize(Decimal('0.01'))

        # Monthly deductions split by period
        sss        = (calculate_sss(monthly)        / divisor).quantize(Decimal('0.01'))
        pagibig    = (calculate_pagibig(monthly)    / divisor).quantize(Decimal('0.01'))
        philhealth = (calculate_philhealth(monthly) / divisor).quantize(Decimal('0.01'))
        tax        = (calculate_tax(monthly)        / divisor).quantize(Decimal('0.01'))

        loan_record    = emp.employee_loans.first()
        loan_deduction = (
            (loan_record.monthly_deductions / divisor) if loan_record else Decimal('0')
        ).quantize(Decimal('0.01'))

        # LWOP: absent days + late (1 day = 480 working minutes)
        daily_rate = monthly / Decimal('22')
        lwop = (
            Decimal(str(absent_days)) * daily_rate
            + Decimal(str(late_mins)) / Decimal('480') * daily_rate
        ).quantize(Decimal('0.01'))

        total_deductions = sss + pagibig + philhealth + tax + loan_deduction + lwop + adj_deductions
        net_pay          = (period_gross - total_deductions).quantize(Decimal('0.01'))

        emp.period_gross       = period_gross
        emp.days_worked        = days_worked
        emp.total_working_days = total_working_days
        emp.absent_days        = absent_days
        emp.late_minutes       = late_mins
        emp.leave_days         = leave_count
        emp.ot_hours           = ot_hours
        emp.ot_pay             = ot_pay
        emp.thirteenth_month   = thirteenth
        emp.adj_earnings       = adj_earnings
        emp.adj_deductions     = adj_deductions
        emp.total_deductions   = total_deductions
        emp.net_pay            = net_pay

        total_gross          += period_gross
        total_deductions_all += total_deductions
        total_net_all        += net_pay

    # Current period adjustments for the form
    from .models import PayrollAdjustment
    current_adjustments = PayrollAdjustment.objects.filter(
        effective_month=period_start.month,
        effective_year=period_start.year,
    ).select_related('employee')

    context = {
        'employees':            employees,
        'total':                total_gross.quantize(Decimal('0.01')),
        'total_deductions':     total_deductions_all.quantize(Decimal('0.01')),
        'total_net':            total_net_all.quantize(Decimal('0.01')),
        'frequency':            frequency,
        'payroll_period_label': period_label,
        'payroll_period_type':  freq_label,
        'freq_divisor':         int(divisor),
        'current_adjustments':  current_adjustments,
        'period_month':         period_start.month,
        'period_year':          period_start.year,
    }
    return render(request, 'sidebuttons/payroll.html', context)


@staff_required
def payroll_detail(request, emp_id):
    employee = get_object_or_404(Employees, id=emp_id)
    monthly  = Decimal(str(employee.basic_pay))

    sss        = calculate_sss(monthly)
    pagibig    = calculate_pagibig(monthly)
    philhealth = calculate_philhealth(monthly)
    tax        = calculate_tax(monthly)

    loan_record    = employee.employee_loans.first()
    loan_deduction = loan_record.monthly_deductions if loan_record else Decimal('0')

    total_deductions = sss + pagibig + philhealth + tax + loan_deduction
    net_pay          = monthly - total_deductions

    context = {
        'employee':         employee,
        'gross':            monthly,
        'sss':              sss.quantize(Decimal('0.01')),
        'pagibig':          pagibig.quantize(Decimal('0.01')),
        'philhealth':       philhealth.quantize(Decimal('0.01')),
        'tax':              tax.quantize(Decimal('0.01')),
        'loan_deduction':   loan_deduction.quantize(Decimal('0.01')),
        'total_deductions': total_deductions.quantize(Decimal('0.01')),
        'net_pay':          net_pay.quantize(Decimal('0.01')),
        'first_period':     (net_pay / 2).quantize(Decimal('0.01')),
        'second_period':    (net_pay / 2).quantize(Decimal('0.01')),
    }
    return render(request, 'sidebuttons/payroll_detail.html', context)


# ── Payroll calculation helpers ───────────────────────────────────────────────

def calculate_tax(monthly_pay):
    m = Decimal(str(monthly_pay))
    if m <= 20833:
        return Decimal('0')
    elif m <= 33333:
        return (m - 20833) * Decimal('0.15')
    elif m <= 66667:
        return Decimal('1875') + (m - 33333) * Decimal('0.20')
    elif m <= 166667:
        return Decimal('8541.80') + (m - 66667) * Decimal('0.25')
    elif m <= 666667:
        return Decimal('33541.80') + (m - 166667) * Decimal('0.30')
    else:
        return Decimal('183541.80') + (m - 666667) * Decimal('0.35')


def calculate_pagibig(monthly_pay):
    m = Decimal(str(monthly_pay))
    return min(m * Decimal('0.02'), Decimal('200'))


def calculate_sss(monthly_pay):
    m = Decimal(str(monthly_pay))
    return m * Decimal('0.05')


def calculate_philhealth(monthly_pay):
    m = Decimal(str(monthly_pay))
    return (m * Decimal('0.05')) / 2


def late_tracker(time_in_dt, duty_in_time):
    """Return minutes late as int. Returns 0 if on time or early."""
    local_dt    = timezone.localtime(time_in_dt)
    duty_dt     = datetime.datetime.combine(local_dt.date(), duty_in_time)
    local_naive = local_dt.replace(tzinfo=None)
    if local_naive > duty_dt:
        return int((local_naive - duty_dt).total_seconds() // 60)
    return 0


def calculate_overtime(attendance_logs, duty_out_time, basic_pay):
    """Return (total_ot_hours as Decimal, ot_pay as Decimal)."""
    hourly_rate = Decimal(str(basic_pay)) / Decimal('22') / Decimal('8')
    ot_rate     = hourly_rate * Decimal('1.25')
    total_ot    = Decimal('0')

    for att in attendance_logs:
        if att.time_out:
            local_out = timezone.localtime(att.time_out)
            duty_dt   = datetime.datetime.combine(local_out.date(), duty_out_time)
            diff_secs = (local_out.replace(tzinfo=None) - duty_dt).total_seconds()
            if diff_secs > 0:
                total_ot += Decimal(str(diff_secs)) / Decimal('3600')

    total_ot = total_ot.quantize(Decimal('0.01'))
    return total_ot, (total_ot * ot_rate).quantize(Decimal('0.01'))


def calculate_13th_month(basic_pay, date_hired):
    """Return accrued 13th month pay as of today."""
    today = datetime.date.today()
    if date_hired.year == today.year:
        months_worked = today.month - date_hired.month + 1
    else:
        months_worked = today.month
    months_worked = max(months_worked, 0)
    return (Decimal(str(basic_pay)) * Decimal(str(months_worked)) / Decimal('12')).quantize(Decimal('0.01'))


def get_adjustment_totals(employee, period_start):
    """Return (earnings_total, deductions_total) for adjustment in the period month."""
    from .models import PayrollAdjustment
    adjs = PayrollAdjustment.objects.filter(
        employee=employee,
        effective_month=period_start.month,
        effective_year=period_start.year,
    )
    earnings = sum(a.amount for a in adjs if a.is_earning)
    deductions = sum(a.amount for a in adjs if not a.is_earning)
    return Decimal(str(earnings)), Decimal(str(deductions))


# ── Excel payroll export ─────────────────────────────────────────────────────

@staff_required
def download_payroll(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter

    today = datetime.date.today()
    last_day = cal.monthrange(today.year, today.month)[1]
    period_start = today.replace(day=1)
    period_end   = today.replace(day=last_day)
    period_str   = f"{period_start.strftime('%B %d, %Y')}  -  {period_end.strftime('%B %d, %Y')}"

    employees_qs = Employees.objects.prefetch_related('employee_loans').all()

    # ── Build per-employee payroll rows ─────────────────────────────────
    rows = []
    for emp in employees_qs:
        m = Decimal(str(emp.basic_pay))

        sss_personal   = (m * Decimal('0.05')).quantize(Decimal('0.01'))
        sss_employer   = (m * Decimal('0.10')).quantize(Decimal('0.01'))
        sss_ec         = Decimal('30')
        pagibig_ee     = calculate_pagibig(m).quantize(Decimal('0.01'))
        pagibig_er     = Decimal('200')
        phil_personal  = calculate_philhealth(m).quantize(Decimal('0.01'))
        phil_employer  = phil_personal
        coop_share     = Decimal('300')

        # Taxable income: salary + 13th-month excess over 90k (monthly share) - mandatory EE deductions
        bonus_excess = max(m - Decimal('90000'), Decimal('0')) / Decimal('12')
        taxable = m + bonus_excess - sss_personal - pagibig_ee - phil_personal

        tax = calculate_tax(taxable).quantize(Decimal('0.01'))

        loan_record = emp.employee_loans.first()

        def lv(field):
            if loan_record:
                v = getattr(loan_record, field, None)
                return Decimal(str(v)) if v else Decimal('0')
            return Decimal('0')

        sss_salary_ln   = lv('SSS_salary_monthly')
        sss_calamity_ln = lv('SSS_calamity_monthly')
        sss_mpl_ln      = lv('SSS_MPL_monthly')
        sss_educ_ln     = lv('SSS_educ_monthly')
        pagibig_housing = lv('PAGIBIG_housing_monthly')
        pagibig_mpl     = lv('PAGIBIG_MPL_monthly')
        coop_loan       = lv('COOP_monthly')

        lwop = Decimal('0')

        total_deductions = (
            tax + sss_personal
            + sss_salary_ln + sss_calamity_ln + sss_mpl_ln + sss_educ_ln
            + pagibig_ee + pagibig_mpl + pagibig_housing
            + phil_personal + coop_share + coop_loan + lwop
        )
        net_pay = (m - total_deductions).quantize(Decimal('0.01'))
        first_half  = (net_pay / 2).quantize(Decimal('0.01'))
        second_half = (net_pay / 2).quantize(Decimal('0.01'))

        rows.append({
            'last_name':    emp.last_name.upper(),
            'first_name':   emp.first_name.upper(),
            'position':     emp.designation,
            'salary':       float(m),
            'taxable':      float(taxable),
            'tax':          float(tax),
            'sss_ee':       float(sss_personal),
            'sss_er':       float(sss_employer),
            'sss_ec':       float(sss_ec),
            'sss_salary':   float(sss_salary_ln),
            'sss_calamity': float(sss_calamity_ln),
            'sss_mpl':      float(sss_mpl_ln),
            'sss_educ':     float(sss_educ_ln),
            'pagibig_ee':   float(pagibig_ee),
            'pagibig_er':   float(pagibig_er),
            'pagibig_mp2':  0.0,
            'pagibig_housing': float(pagibig_housing),
            'pagibig_mpl':  float(pagibig_mpl),
            'phil_ee':      float(phil_personal),
            'phil_er':      float(phil_employer),
            'coop_share':   float(coop_share),
            'coop_loan':    float(coop_loan),
            'lwop':         float(lwop),
            'net_pay':      float(net_pay),
            'first':        float(first_half),
            'second':       float(second_half),
        })

    # ── Build workbook ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payroll'

    # Styles
    bold   = Font(bold=True)
    bold_l = Font(bold=True, size=14)
    header_font = Font(bold=True, size=9)
    data_font   = Font(size=10)
    money_fmt   = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_a = Alignment(vertical='center', wrap_text=True)

    # Column widths
    col_widths = {
        'A': 5, 'B': 14, 'C': 18, 'D': 16, 'E': 14, 'F': 14, 'G': 14,
        'H': 12, 'I': 12, 'J': 8,
        'K': 12, 'L': 12, 'M': 12, 'N': 12,
        'O': 12, 'P': 12, 'Q': 12, 'R': 12, 'S': 12,
        'T': 12, 'U': 12,
        'V': 12, 'W': 12,
        'X': 14,
        'Y': 3,
        'Z': 14, 'AA': 14, 'AB': 14,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Row 2: Company name ─────────────────────────────────────────────
    ws.merge_cells('B2:D2')
    ws['B2'] = 'PAYROLL REGISTER'
    ws['B2'].font = bold_l

    # ── Row 3: Period ───────────────────────────────────────────────────
    ws.merge_cells('B3:E3')
    ws['B3'] = f'PERIOD : {period_str}'
    ws['B3'].font = Font(bold=True, size=10)

    # ── Row 6-7: Headers ────────────────────────────────────────────────
    # Row 6: main headers (some merged across sub-columns)
    main_headers = [
        ('A6', 'A7', 'NO'),
        ('B6', 'C7', 'NAME'),
        ('D6', 'D7', 'POSITION'),
        ('E6', 'E7', 'SALARY'),
        ('F6', 'F7', 'Taxable\nIncome'),
    ]

    for start, end, label in main_headers:
        ws.merge_cells(f'{start}:{end}')
        ws[start] = label

    # WITHHOLDING TAX
    ws['G6'] = 'WITHHOLDING'
    ws['G7'] = 'TAX'

    # SSS CONTRIBUTIONS (H-J header, sub-headers in row 7)
    ws.merge_cells('H6:J6')
    ws['H6'] = 'SSS CONTRIBUTIONS'
    ws['H7'] = 'Personal'
    ws['I7'] = 'Employer'
    ws['J7'] = 'EC'

    # SSS LOANS (K-N)
    ws.merge_cells('K6:N6')
    ws['K6'] = 'SSS LOANS'
    ws['K7'] = 'SALARY'
    ws['L7'] = 'CALAMITY'
    ws['M7'] = 'MPL'
    ws['N7'] = 'EDUC'

    # PAGIBIG (O-Q)
    ws.merge_cells('O6:Q6')
    ws['O6'] = 'PAGIBIG'
    ws['O7'] = 'PERSONAL'
    ws['P7'] = 'EMPLOYER'
    ws['Q7'] = 'PAG-IBIG MP2'

    # PAGIBIG LOAN (R-S)
    ws.merge_cells('R6:S6')
    ws['R6'] = 'PAGIBIG LOAN'
    ws['R7'] = 'HOUSING'
    ws['S7'] = 'MPL'

    # PHILHEALTH (T-U)
    ws.merge_cells('T6:U6')
    ws['T6'] = 'PHILHEALTH'
    ws['T7'] = 'PERSONAL'
    ws['U7'] = 'EMPLOYER'

    # COOPERATIVE (V-W)
    ws.merge_cells('V6:W6')
    ws['V6'] = 'COOPERATIVE'
    ws['V7'] = 'CONTRIBUTION'
    ws['W7'] = 'LOAN'

    # LWOP
    ws['X6'] = 'LEAVE WITHOUT'
    ws['X7'] = 'PAY'

    # NET PAY / PERIODS
    ws['Z6'] = 'NET PAY'
    ws['AA6'] = '1ST PERIOD'
    ws['AB6'] = '2ND PERIOD'

    # Style header rows
    for row_num in (6, 7):
        for col_num in range(1, 29):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
            cell.fill = header_fill

    # ── Data rows (starting row 9) ─────────────────────────────────────
    data_start = 9
    col_map = [
        None,            # A = row number (set manually)
        'last_name',     # B
        'first_name',    # C
        'position',      # D
        'salary',        # E
        'taxable',       # F
        'tax',           # G
        'sss_ee',        # H
        'sss_er',        # I
        'sss_ec',        # J
        'sss_salary',    # K
        'sss_calamity',  # L
        'sss_mpl',       # M
        'sss_educ',      # N
        'pagibig_ee',    # O
        'pagibig_er',    # P
        'pagibig_mp2',   # Q
        'pagibig_housing',  # R
        'pagibig_mpl',   # S
        'phil_ee',       # T
        'phil_er',       # U
        'coop_share',    # V
        'coop_loan',     # W
        'lwop',          # X
        None,            # Y (spacer)
        'net_pay',       # Z
        'first',         # AA
        'second',        # AB
    ]

    for i, row_data in enumerate(rows):
        r = data_start + i
        for c, key in enumerate(col_map, start=1):
            cell = ws.cell(row=r, column=c)
            if c == 1:
                cell.value = i + 1
                cell.alignment = center
            elif key is None:
                pass
            elif key in ('last_name', 'first_name', 'position'):
                cell.value = row_data[key]
                cell.alignment = left_a
            else:
                val = row_data[key]
                cell.value = val if val else None
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal='right', vertical='center')

            cell.font = data_font
            cell.border = thin_border

    # ── Totals row ──────────────────────────────────────────────────────
    totals_row = data_start + len(rows) + 1
    ws.merge_cells(f'B{totals_row}:D{totals_row}')
    ws[f'B{totals_row}'] = 'TOTAL OR CARRIED FORWARD'
    ws[f'B{totals_row}'].font = Font(bold=True, size=10)

    sum_cols = [5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 26, 27, 28]
    for c in sum_cols:
        cell = ws.cell(row=totals_row, column=c)
        col_letter = get_column_letter(c)
        cell.value = f'=SUM({col_letter}{data_start}:{col_letter}{totals_row - 2})'
        cell.number_format = money_fmt
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border

    # ── Signature block ─────────────────────────────────────────────────
    sig_row = totals_row + 3
    ws[f'B{sig_row}'] = 'CERTIFIED BY:'
    ws[f'B{sig_row}'].font = bold
    ws[f'G{sig_row}'] = 'AVAILABILITY OF FUNDS:'
    ws[f'G{sig_row}'].font = bold

    ws[f'C{sig_row + 2}'] = 'Department Head'
    ws[f'C{sig_row + 2}'].font = Font(italic=True, size=10)

    # ── Return as download ──────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Payroll_{period_start.strftime('%b_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── Excel payslip export ─────────────────────────────────────────────────────

@login_required
def download_payslip(request, emp_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    employee = get_object_or_404(Employees, id=emp_id)
    m = Decimal(str(employee.basic_pay))

    today        = datetime.date.today()
    last_day     = cal.monthrange(today.year, today.month)[1]
    period_start = today.replace(day=1)
    period_end   = today.replace(day=last_day)
    period_str   = f"{period_start.strftime('%B %d')} - {period_end.strftime('%B %d, %Y')}"

    # Compute deductions
    sss_ee      = (m * Decimal('0.05')).quantize(Decimal('0.01'))
    sss_er      = (m * Decimal('0.10')).quantize(Decimal('0.01'))
    sss_ec      = Decimal('30')
    pagibig_ee  = calculate_pagibig(m).quantize(Decimal('0.01'))
    pagibig_er  = Decimal('200')
    phil_ee     = calculate_philhealth(m).quantize(Decimal('0.01'))
    phil_er     = phil_ee
    coop_share  = Decimal('300')

    bonus_excess = max(m - Decimal('90000'), Decimal('0')) / Decimal('12')
    taxable = m + bonus_excess - sss_ee - pagibig_ee - phil_ee
    tax = calculate_tax(taxable).quantize(Decimal('0.01'))

    loan_record = employee.employee_loans.first()

    def lv(field):
        if loan_record:
            v = getattr(loan_record, field, None)
            return Decimal(str(v)) if v else Decimal('0')
        return Decimal('0')

    loan_items = [
        ('SSS Salary Loan',   lv('SSS_salary_monthly')),
        ('SSS Calamity Loan', lv('SSS_calamity_monthly')),
        ('SSS MPL',           lv('SSS_MPL_monthly')),
        ('SSS Educational',   lv('SSS_educ_monthly')),
        ('Pag-IBIG Housing',  lv('PAGIBIG_housing_monthly')),
        ('Pag-IBIG MPL',      lv('PAGIBIG_MPL_monthly')),
        ('Coop Loan',         lv('COOP_monthly')),
    ]
    active_loans = [(name, amt) for name, amt in loan_items if amt > 0]
    total_loans  = sum(amt for _, amt in active_loans)

    total_deductions = (
        tax + sss_ee + pagibig_ee + phil_ee + coop_share + total_loans
    ).quantize(Decimal('0.01'))
    net_pay     = (m - total_deductions).quantize(Decimal('0.01'))
    first_half  = (net_pay / 2).quantize(Decimal('0.01'))
    second_half = (net_pay / 2).quantize(Decimal('0.01'))

    # ── Build workbook ──────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Payslip'

    # Styles
    money_fmt = '#,##0.00'
    thin = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin'),
    )
    blue_fill  = PatternFill('solid', fgColor='1A234E')
    light_fill = PatternFill('solid', fgColor='F0F4F8')
    green_fill = PatternFill('solid', fgColor='F0FDF4')
    bold_w = Font(bold=True, color='FFFFFF', size=11)
    bold_b = Font(bold=True, size=10)
    normal = Font(size=10)
    label_f = Font(size=9, color='6B7280')
    earn_f  = Font(size=10, color='10B981', bold=True)
    ded_f   = Font(size=10, color='EF4444', bold=True)
    net_f   = Font(size=14, color='0369A1', bold=True)
    right_a = Alignment(horizontal='right', vertical='center')
    left_a  = Alignment(horizontal='left', vertical='center')
    center  = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 3

    def styled_row(row, label, value, label_font=label_f, val_font=normal,
                   val_fmt=money_fmt, fill=None):
        ws.cell(row=row, column=2, value=label).font = label_font
        ws.cell(row=row, column=2).alignment = left_a
        c = ws.cell(row=row, column=4, value=float(value) if isinstance(value, Decimal) else value)
        c.font = val_font
        c.number_format = val_fmt
        c.alignment = right_a
        if fill:
            for col in range(2, 5):
                ws.cell(row=row, column=col).fill = fill

    # ── Header bar ──────────────────────────────────────────────────────
    ws.merge_cells('B2:D2')
    h = ws['B2']
    h.value = 'PAYSLIP'
    h.font  = bold_w
    h.fill  = blue_fill
    h.alignment = center
    for c in range(2, 5):
        ws.cell(row=2, column=c).fill = blue_fill

    # Employee info
    ws.cell(row=4, column=2, value='Employee').font = label_f
    ws.cell(row=4, column=3, value=f'{employee.first_name} {employee.last_name}').font = bold_b

    ws.cell(row=5, column=2, value='Designation').font = label_f
    ws.cell(row=5, column=3, value=employee.designation).font = normal

    ws.cell(row=6, column=2, value='Employee ID').font = label_f
    ws.cell(row=6, column=3, value=employee.employee_id).font = normal

    ws.cell(row=7, column=2, value='Pay Period').font = label_f
    ws.cell(row=7, column=3, value=period_str).font = normal

    # ── Earnings ────────────────────────────────────────────────────────
    r = 9
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(row=r, column=2, value='EARNINGS').font = Font(bold=True, size=9, color='6B7280')
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = light_fill
        ws.cell(row=r, column=c).border = thin

    r = 10
    styled_row(r, 'Monthly Salary', m, val_font=bold_b)
    r = 11
    styled_row(r, 'Taxable Income', taxable)
    r = 12
    styled_row(r, 'Period Gross (1/2)', m / 2, val_font=earn_f)

    # ── Statutory Deductions ────────────────────────────────────────────
    r = 14
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(row=r, column=2, value='STATUTORY DEDUCTIONS').font = Font(bold=True, size=9, color='6B7280')
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = light_fill
        ws.cell(row=r, column=c).border = thin

    deductions = [
        ('Withholding Tax',         tax),
        ('SSS (Employee 5%)',       sss_ee),
        ('SSS (Employer 10%)',      sss_er),
        ('SSS EC',                  sss_ec),
        ('Pag-IBIG (Employee)',     pagibig_ee),
        ('Pag-IBIG (Employer)',     pagibig_er),
        ('PhilHealth (Employee)',   phil_ee),
        ('PhilHealth (Employer)',   phil_er),
        ('Coop Contribution',      coop_share),
    ]

    r = 15
    for label, amt in deductions:
        is_employer = 'Employer' in label
        styled_row(r, label, amt, val_font=Font(size=10, color='9CA3AF' if is_employer else 'EF4444'))
        r += 1

    # ── Loan Deductions ─────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(row=r, column=2, value='LOAN DEDUCTIONS').font = Font(bold=True, size=9, color='6B7280')
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = light_fill
        ws.cell(row=r, column=c).border = thin
    r += 1

    if active_loans:
        for label, amt in active_loans:
            styled_row(r, f'  {label}', amt, val_font=ded_f)
            r += 1
        styled_row(r, 'Total Loan Deductions', total_loans, label_font=bold_b, val_font=ded_f)
        r += 1
    else:
        ws.cell(row=r, column=2, value='  No active loans').font = Font(size=9, color='9CA3AF', italic=True)
        r += 1

    # ── Summary ─────────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f'B{r}:D{r}')
    ws.cell(row=r, column=2, value='SUMMARY').font = Font(bold=True, size=9, color='6B7280')
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = light_fill
        ws.cell(row=r, column=c).border = thin
    r += 1

    styled_row(r, 'Gross Pay', m, val_font=bold_b)
    r += 1
    styled_row(r, 'Total Deductions', total_deductions, val_font=ded_f)
    r += 1

    # Net pay highlight
    r += 1
    for c in range(2, 5):
        ws.cell(row=r, column=c).fill = green_fill
        ws.cell(row=r, column=c).border = thin
    ws.cell(row=r, column=2, value='NET PAY').font = Font(bold=True, size=12, color='0369A1')
    ws.cell(row=r, column=2).alignment = left_a
    c = ws.cell(row=r, column=4, value=float(net_pay))
    c.font = net_f
    c.number_format = money_fmt
    c.alignment = right_a
    ws.row_dimensions[r].height = 30

    # Period split
    r += 2
    styled_row(r, '1st Period (1st - 15th)', first_half, val_font=bold_b)
    r += 1
    styled_row(r, '2nd Period (16th - End)', second_half, val_font=bold_b)

    # ── Return as download ──────────────────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_name = f"{employee.last_name}_{employee.first_name}".replace(' ', '_')
    filename = f"Payslip_{safe_name}_{today.strftime('%b_%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── Payroll History ──────────────────────────────────────────────────────────

@staff_required
def lock_payroll(request):
    if request.method != 'POST':
        return redirect('SIDEBUTTON-PAYROLL')

    frequency = request.POST.get('frequency', 'monthly')
    today     = datetime.date.today()

    if frequency == 'weekly':
        period_start = today - datetime.timedelta(days=today.weekday())
        period_end   = period_start + datetime.timedelta(days=6)
        divisor      = Decimal('4')
    elif frequency == 'biweekly':
        if today.day <= 15:
            period_start = today.replace(day=1)
            period_end   = today.replace(day=15)
        else:
            last_day     = cal.monthrange(today.year, today.month)[1]
            period_start = today.replace(day=16)
            period_end   = today.replace(day=last_day)
        divisor = Decimal('2')
    else:
        period_start = today.replace(day=1)
        last_day     = cal.monthrange(today.year, today.month)[1]
        period_end   = today.replace(day=last_day)
        divisor      = Decimal('1')

    if PayrollRun.objects.filter(start_date=period_start, end_date=period_end, frequency=frequency).exists():
        messages.error(request, 'Payroll for this period has already been locked.')
        return redirect('SIDEBUTTON-PAYROLL')

    effective_end      = min(period_end, today)
    total_working_days = sum(
        1 for i in range((effective_end - period_start).days + 1)
        if (period_start + datetime.timedelta(days=i)).weekday() < 5
    )

    attendance_by_emp = {}
    for att in Attendance.objects.filter(date__range=(period_start, effective_end)).select_related('employee'):
        attendance_by_emp.setdefault(att.employee_id, []).append(att)

    employees    = list(Employees.objects.filter(is_archive=False).prefetch_related('employee_loans'))
    run_gross    = Decimal('0')
    run_ded      = Decimal('0')
    run_net      = Decimal('0')

    run = PayrollRun.objects.create(
        start_date=period_start, end_date=period_end, frequency=frequency,
        created_by=request.user,
    )

    for emp in employees:
        m           = Decimal(str(emp.basic_pay))
        period_gross = (m / divisor).quantize(Decimal('0.01'))

        emp_logs    = attendance_by_emp.get(emp.id, [])
        days_worked = len(emp_logs)
        absent_days = max(0, total_working_days - days_worked)
        late_mins   = sum(late_tracker(a.time_in, emp.duty_in) for a in emp_logs if a.time_in)
        leave_count = LeaveRequest.objects.filter(
            employee=emp, status='APPROVED',
            start_date__lte=effective_end, end_date__gte=period_start,
        ).count()

        ot_hours, ot_pay = calculate_overtime(emp_logs, emp.duty_out, m)
        adj_earnings, adj_deductions = get_adjustment_totals(emp, period_start)
        thirteenth = calculate_13th_month(m, emp.date_hired)

        period_gross = (m / divisor + ot_pay + adj_earnings).quantize(Decimal('0.01'))

        sss        = (calculate_sss(m) / divisor).quantize(Decimal('0.01'))
        pagibig    = (calculate_pagibig(m) / divisor).quantize(Decimal('0.01'))
        philhealth = (calculate_philhealth(m) / divisor).quantize(Decimal('0.01'))
        tax        = (calculate_tax(m) / divisor).quantize(Decimal('0.01'))

        loan_record    = emp.employee_loans.first()
        loan_deduction = ((loan_record.monthly_deductions / divisor) if loan_record else Decimal('0')).quantize(Decimal('0.01'))

        daily_rate = m / Decimal('22')
        lwop = (Decimal(str(absent_days)) * daily_rate + Decimal(str(late_mins)) / Decimal('480') * daily_rate).quantize(Decimal('0.01'))

        total_ded = (sss + pagibig + philhealth + tax + loan_deduction + lwop + adj_deductions).quantize(Decimal('0.01'))
        net_pay   = (period_gross - total_ded).quantize(Decimal('0.01'))

        pp = PayrollPeriod.objects.create(
            payroll_run=run, employee=emp,
            start_date=period_start, end_date=period_end,
            days_worked=days_worked, absent_days=absent_days,
            late_minutes=late_mins, leave_days=leave_count,
            gross_salary=period_gross, net_salary=net_pay,
            loan_deductions=loan_deduction, lwop_amount=lwop,
            ot_hours=ot_hours, ot_pay=ot_pay,
            thirteenth_month=thirteenth,
            adjustment_total=adj_earnings - adj_deductions,
            first_period_pay=(net_pay / 2).quantize(Decimal('0.01')),
            second_period_pay=(net_pay / 2).quantize(Decimal('0.01')),
        )

        Contribution.objects.create(
            employee=emp, payroll=pp,
            sss_personal=sss,
            sss_employer=(m * Decimal('0.10') / divisor).quantize(Decimal('0.01')),
            pagibig_personal=pagibig, philhealth_personal=philhealth,
            philhealth_employer=philhealth, tax=tax,
        )

        run_gross += period_gross
        run_ded   += total_ded
        run_net   += net_pay

    run.employee_count   = len(employees)
    run.total_gross      = run_gross.quantize(Decimal('0.01'))
    run.total_deductions = run_ded.quantize(Decimal('0.01'))
    run.total_net        = run_net.quantize(Decimal('0.01'))
    run.save()

    log_action(request, 'PAYROLL_LOCKED', f'Locked {frequency} payroll for {period_start.strftime("%b %d")} - {period_end.strftime("%b %d, %Y")} ({len(employees)} employees).')
    messages.success(request, f'Payroll locked for {period_start.strftime("%b %d")} - {period_end.strftime("%b %d, %Y")}.')
    return redirect('PAYROLL-HISTORY-DETAIL', run_id=run.id)


@staff_required
def payroll_history(request):
    runs = PayrollRun.objects.all()
    return render(request, 'sidebuttons/payroll_history.html', {'runs': runs})


@staff_required
def payroll_history_detail(request, run_id):
    run     = get_object_or_404(PayrollRun, id=run_id)
    periods = run.periods.select_related('employee').all()
    return render(request, 'sidebuttons/payroll_history_detail.html', {'run': run, 'periods': periods})


@staff_required
def add_adjustment(request):
    if request.method == 'POST':
        from .models import PayrollAdjustment

        emp_id  = request.POST.get('employee_id')
        adj_type = request.POST.get('adjustment_type')
        amount  = request.POST.get('amount')
        desc    = request.POST.get('description', '')
        month   = int(request.POST.get('effective_month', datetime.date.today().month))
        year    = int(request.POST.get('effective_year', datetime.date.today().year))

        employee = get_object_or_404(Employees, id=emp_id)

        PayrollAdjustment.objects.create(
            employee=employee,
            adjustment_type=adj_type,
            amount=Decimal(amount),
            description=desc,
            effective_month=month,
            effective_year=year,
            created_by=request.user,
        )
        log_action(request, 'ADJUSTMENT_ADDED', f'{adj_type.replace("_", " ").title()} of P{amount} for {employee.first_name} {employee.last_name}: {desc}')
        messages.success(request, f'{adj_type.replace("_", " ").title()} of ₱{amount} added for {employee.first_name} {employee.last_name}.')

    return redirect('SIDEBUTTON-PAYROLL')
